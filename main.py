from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from service.pdf_generator import build_pdf

REQUIRED_FIELDS = (
    "INVOICE_NUMBER",
    "DATE",
    "DUE_DATE",
    "ITEM_DESCRIPTION",
    "QTY",
    "UNIT_PRICE",
    "GST_PERCENT",
)

COLUMN_ALIASES = {
    "COMPANY NAME": "COMPANY_NAME",
    "COMPANY ID": "COMPANY_ID",
    "INVOICE NUMBER": "INVOICE_NUMBER",
    "DATE": "DATE",
    "DUE DATE": "DUE_DATE",
    "ADDRESS": "ADDRESS",
    "ITEM DESCRIPTION": "ITEM_DESCRIPTION",
    "QTY": "QTY",
    "UNIT PRICE": "UNIT_PRICE",
    "LINE TOTAL": "LINE_TOTAL",
    "GST %": "GST_PERCENT",
    "AMOUNT": "AMOUNT",
}


def normalize_headers(headers: list[str]) -> list[str]:
    return [COLUMN_ALIASES.get(h.strip(), h.strip()) for h in headers]


def parse_args() -> argparse.Namespace:
    base_dir = Path(__file__).resolve().parent
    input_dir = base_dir / "input"

    default_excel = input_dir / "source_invoice_data.xlsx"
    default_csv = input_dir / "invoice_data.csv"

    if default_excel.exists():
        default_source = "excel"
        default_input = default_excel
    else:
        default_source = "csv"
        default_input = default_csv

    parser = argparse.ArgumentParser(description="Generate invoice PDFs from CSV, Excel, or interactive input.")
    parser.add_argument("--source", choices=("csv", "excel", "cli"), default=default_source, help="Input source for invoices.")
    parser.add_argument("--config", type=Path, default=base_dir / "configuration" / "config.json", help="Path to the configuration JSON file.")
    parser.add_argument("--input", type=Path, default=default_input, help="Path to the input CSV or Excel file.")
    parser.add_argument("--output", type=Path, default=base_dir / "output", help="Directory where generated PDFs are written.")
    return parser.parse_args()


def prompt_input(prompt: str, optional: bool = False) -> str:
    while True:
        value = input(prompt).strip()
        if value or optional:
            return value
        print("  Input cannot be blank. Please try again.")


def prompt_number(prompt: str, value_type: type[float] | type[int]) -> str:
    while True:
        value = prompt_input(prompt)
        try:
            parsed = value_type(value)
        except ValueError:
            print("  Please enter a valid number.")
            continue

        if parsed < 0:
            print("  Please enter a non-negative value.")
            continue

        return str(parsed)


def prompt_for_invoice_record() -> dict[str, str]:
    print("\nEnter invoice details:")
    return {
        "INVOICE_NUMBER": prompt_input("  Invoice number: "),
        "DATE": prompt_input("  Issue date (dd/mm/yyyy): "),
        "DUE_DATE": prompt_input("  Due date (dd/mm/yyyy): "),
        "COMPANY_NAME": prompt_input("  Company name (leave blank for default): ", optional=True),
        "COMPANY_ID": prompt_input("  Company ID (leave blank for default): ", optional=True),
        "ADDRESS": prompt_input("  Billing address (leave blank for default): ", optional=True),
        "ITEM_DESCRIPTION": prompt_input("  Item description: "),
        "QTY": prompt_number("  Quantity: ", int),
        "UNIT_PRICE": prompt_number("  Unit price: ", float),
        "GST_PERCENT": prompt_number("  GST percent: ", float),
    }


def read_invoice_records_from_csv(file_path: Path) -> list[dict[str, str]]:
    if not file_path.exists():
        raise FileNotFoundError(f"CSV input file not found: {file_path}")

    with file_path.open("r", encoding="utf-8", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        raw_fieldnames = reader.fieldnames or []
        fieldnames = normalize_headers(raw_fieldnames)
        reader.fieldnames = fieldnames
        missing_columns = [field for field in REQUIRED_FIELDS if field not in fieldnames]
        if missing_columns:
            raise ValueError(f"CSV file is missing required columns: {', '.join(missing_columns)}")
        records = [row for row in reader if any(row.values())]

    if not records:
        raise ValueError("No invoice records found in CSV input.")

    return records


def read_invoice_records_from_excel(file_path: Path) -> list[dict[str, str]]:
    if not file_path.exists():
        raise FileNotFoundError(f"Excel input file not found: {file_path}")

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("No invoice records found in Excel input.")

    raw_headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    headers = normalize_headers(raw_headers)
    missing_columns = [field for field in REQUIRED_FIELDS if field not in headers]
    if missing_columns:
        raise ValueError(f"Excel file is missing required columns: {', '.join(missing_columns)}")

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        record = {
            headers[index]: (str(row[index]) if index < len(row) else "")
            for index in range(len(headers))
        }
        records.append(record)

    if not records:
        raise ValueError("No invoice records found in Excel input.")
    return records


def group_invoice_rows(records: list[dict[str, str]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for record in records:
        invoice_number = str(record.get("INVOICE_NUMBER", "")).strip()
        if not invoice_number:
            continue

        if invoice_number not in grouped:
            grouped[invoice_number] = {
                "INVOICE_NUMBER": invoice_number,
                "DATE": record.get("DATE", ""),
                "DUE_DATE": record.get("DUE_DATE", ""),
                "COMPANY_NAME": record.get("COMPANY_NAME", ""),
                "COMPANY_ID": record.get("COMPANY_ID", ""),
                "ADDRESS": record.get("ADDRESS", ""),
                "ITEMS": [],
            }

        grouped[invoice_number]["ITEMS"].append(
            {
                "ITEM_DESCRIPTION": record.get("ITEM_DESCRIPTION", ""),
                "QTY": record.get("QTY", ""),
                "UNIT_PRICE": record.get("UNIT_PRICE", ""),
                "LINE_TOTAL": record.get("LINE_TOTAL", ""),
                "GST_PERCENT": record.get("GST_PERCENT", ""),
                "AMOUNT": record.get("AMOUNT", ""),
            }
        )

    return list(grouped.values())


def read_invoice_records_from_input(file_path: Path) -> list[dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        records = read_invoice_records_from_excel(file_path)
        print(f"[SUCCESS] Parsed {len(records)} invoice record(s) from Excel.")
    else:
        records = read_invoice_records_from_csv(file_path)
        print(f"[SUCCESS] Parsed {len(records)} invoice record(s) from CSV.")
    return group_invoice_rows(records)


def validate_grouped_record(record: dict[str, Any]) -> None:
    missing_values = [field for field in ("INVOICE_NUMBER", "DATE", "DUE_DATE") if not record.get(field)]
    if missing_values:
        raise ValueError("Invoice group is missing required values: " + ", ".join(missing_values))

    if not record.get("ITEMS"):
        raise ValueError("Invoice group does not contain any line items.")


def process_invoice_records(config_data: dict[str, Any], records: list[dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for index, record in enumerate(records, start=1):
        validate_grouped_record(record)
        print(f" -> Mapping record #{index}: Invoice {record['INVOICE_NUMBER']}")
        build_pdf(config_data, record, output_dir)


def execute_autonomous_pipeline() -> None:
    args = parse_args()
    print("--- Running Native Python Backend Pipeline ---")

    with args.config.open("r", encoding="utf-8") as handle:
        config_data = json.load(handle)
    print("[SUCCESS] Backend loaded JSON rules database configurations.")

    if args.source == "cli":
        records: list[dict[str, str]] = []
        while True:
            records.append(prompt_for_invoice_record())
            more = prompt_input("Add another invoice? [y/N]: ", optional=True).lower()
            if more not in ("y", "yes"):
                break
        grouped_records = group_invoice_rows(records)
    else:
        grouped_records = read_invoice_records_from_input(args.input)

    process_invoice_records(config_data, grouped_records, args.output)
    print(f"\n[PIPELINE EXITED SUCCESSFULLY] Documents routed out directly to: {args.output.resolve()}")


if __name__ == "__main__":
    execute_autonomous_pipeline()